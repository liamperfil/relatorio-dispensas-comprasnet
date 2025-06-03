from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import os
from datetime import datetime
from bs4 import BeautifulSoup
import openpyxl
import shutil

def iniciar_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    return webdriver.Chrome(options=options)

def log(mensagem, arquivo_log):
    """Logs messages to console and a specified file."""
    print(mensagem)
    log_dir = os.path.dirname(arquivo_log)
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
    with open(arquivo_log, 'a', encoding='utf-8') as f:
        f.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} - {mensagem}\n')

def salvar_html(driver, pagina, html_dir):
    """Saves the current page HTML to a file in the specified directory."""
    if not os.path.exists(html_dir):
        os.makedirs(html_dir)
        log(f"Diretório criado: {html_dir}", log_filename)
    
    data_atual = datetime.now().strftime("%Y-%m-%d")
    nome_arquivo = os.path.join(html_dir, f'{data_atual}-pg{pagina}.html')
    
    with open(nome_arquivo, 'w', encoding='utf-8') as f:
        f.write(driver.page_source)
    log(f"HTML salvo em {nome_arquivo}", log_filename)

def normalizar_valor_numerico(texto):
    """Normaliza uma string para um float, removendo 'R$', pontos e vírgulas."""
    if not texto:
        return 0.0
    # Remove 'R$', pontos e substitui vírgula por ponto
    texto_limpo = texto.replace('R$', '').replace('.', '').replace(',', '.').strip()
    try:
        return float(texto_limpo)
    except ValueError:
        return 0.0 # Retorna 0.0 se a conversão falhar

def normalizar_qtde(texto):
    """Normaliza a quantidade para um float, removendo pontos e vírgulas."""
    if not texto:
        return 0.0
    # Remove pontos e vírgulas, tenta converter para float
    texto_limpo = texto.replace('.', '').replace(',', '.').strip()
    try:
        return float(texto_limpo)
    except ValueError:
        return 0.0 # Retorna 0.0 se a conversão falhar

def extrair_dados_html(html_content, log_file):
    """
    Extracts item data from the HTML content, specifically looking for
    child tables within each dispense result and all their items.
    Applies data normalization for numerical and monetary fields.
    """
    soup = BeautifulSoup(html_content, 'html.parser')
    dados_extraidos = []

    main_dispense_tables = soup.find_all('table', id='tblResultadoLista')

    for main_table in main_dispense_tables:
        numero_dispensa = ''
        data_abertura = ''
        main_table_body_row = main_table.find('tbody').find('tr', recursive=False)
        if main_table_body_row:
            cols_main = main_table_body_row.find_all('td', recursive=False)
            if len(cols_main) > 0 and cols_main[0].find('a'):
                numero_dispensa = cols_main[0].find('a').get_text(strip=True)
            if len(cols_main) > 1:
                data_abertura = cols_main[1].get_text(strip=True)


        child_table = main_table.find('table', id='tblResultadoLista_Child')
        
        if child_table:
            all_tbodies = child_table.find_all('tbody', recursive=False)

            for tbody in all_tbodies:
                item_row = None
                situacao_item = ''

                for row in tbody.find_all('tr', recursive=False):
                    cols = row.find_all('td', recursive=False)
                    if len(cols) == 7 and not cols[0].has_attr('colspan'):
                        item_row = row
                    elif len(cols) == 1 and cols[0].has_attr('colspan') and 'Situação do Item:' in cols[0].get_text(strip=True):
                        situacao_item = cols[0].get_text(strip=True).replace('Situação do Item:', '').strip()
                
                if item_row:
                    cols = item_row.find_all('td')
                    descricao = cols[0].get_text(strip=True) if len(cols) > 0 and cols[0] else ''
                    uf = cols[1].get_text(strip=True) if len(cols) > 1 and cols[1] else ''
                    vencedor = cols[2].get_text(strip=True) if len(cols) > 2 and cols[2] else ''
                    marca = cols[3].get_text(strip=True) if len(cols) > 3 and cols[3] else ''
                    
                    # Aplica normalização
                    qtde = normalizar_qtde(cols[4].get_text(strip=True)) if len(cols) > 4 and cols[4] else 0.0
                    unitario = normalizar_valor_numerico(cols[5].get_text(strip=True)) if len(cols) > 5 and cols[5] else 0.0
                    total = normalizar_valor_numerico(cols[6].get_text(strip=True)) if len(cols) > 6 and cols[6] else 0.0

                    dados_extraidos.append([numero_dispensa, data_abertura, descricao, uf, vencedor, marca, qtde, unitario, total, situacao_item])
                else:
                    log(f"Nenhuma linha de item com 7 colunas encontrada neste bloco <tbody>. Conteúdo do tbody: {tbody.get_text(strip=True)[:100]}...", log_file)
        else:
            log("Nenhuma tabela 'tblResultadoLista_Child' encontrada em uma tabela principal de dispensa.", log_file)

    return dados_extraidos

def adicionar_dados_a_planilha(dados, nome_planilha='planilha.xlsx'):
    """Adds extracted data to an Excel spreadsheet, creating it with headers if it doesn't exist."""
    if not os.path.exists(nome_planilha):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(['Número', 'Abertura', 'Descrição', 'Uf', 'Vencedor', 'Marca', 'Qtde', 'Unitário', 'Total', 'Situação do Item'])
        log(f"Criado novo arquivo Excel: {nome_planilha} com cabeçalhos.", log_filename)
    else:
        try:
            workbook = openpyxl.load_workbook(nome_planilha)
            sheet = workbook['Sheet1']
            log(f"Abrindo arquivo Excel existente: {nome_planilha}.", log_filename)
        except Exception as e:
            log(f"Erro ao abrir o arquivo Excel {nome_planilha}: {e}. Tentando criar um novo.", log_filename)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(['Número', 'Abertura', 'Descrição', 'Uf', 'Vencedor', 'Marca', 'Qtde', 'Unitário', 'Total', 'Situação do Item'])
            
    for row_data in dados:
        try:
            sheet.append(row_data)
        except Exception as e:
            log(f"Erro ao adicionar linha {row_data} ao Excel: {e}", log_filename)
    workbook.save(nome_planilha)
    log(f"Dados salvos em {nome_planilha}.", log_filename)

def raspar_comprasnet():
    """Automates the web scraping process."""
    driver = iniciar_driver()
    wait = WebDriverWait(driver, 60)
    
    global log_filename
    timestamp = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    log_filename = os.path.join('log', f'log_{timestamp}.txt')

    html_dir = 'html'

    log('Iniciando raspagem no Comprasnet...', log_filename)

    try:
        driver.get("https://comprasnet3.ba.gov.br/CompraEletronica/ResultadoFiltro.asp?token=68385a59c508b")
        log('Site acessado.', log_filename)
        
        log('Aguardando carregamento da tabela de resultados...', log_filename)
        wait.until(EC.visibility_of_element_located((By.ID, 'tblResultadoListaCount')))
        log('Tabela de resultados carregada.', log_filename)

        pagina = 1

        while True:
            log(f'Coletando dados da página {pagina}...', log_filename)

            salvar_html(driver, pagina, html_dir)
            
            html_content = driver.page_source
            dados_extraidos_pagina = extrair_dados_html(html_content, log_filename)
            if dados_extraidos_pagina:
                adicionar_dados_a_planilha(dados_extraidos_pagina)
                log(f'{len(dados_extraidos_pagina)} itens da página {pagina} adicionados à planilha.', log_filename)
            else:
                log(f'Nenhum item encontrado na página {pagina} para adicionar à planilha.', log_filename)

            try:
                botao_proxima = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tblResultadoListaCount_next"]')))
                classe = botao_proxima.get_attribute('class')
                if 'disabled' in classe:
                    log('Não há mais páginas. Finalizando raspagem.', log_filename)
                    break
                else:
                    log(f'Clicando para ir para a próxima página ({pagina + 1})...', log_filename)
                    botao_proxima.click()
                    pagina += 1
                    time.sleep(3)
            except NoSuchElementException:
                log('Botão "Próxima" página não encontrado. Assumindo que não há mais páginas. Finalizando.', log_filename)
                break
            except TimeoutException:
                log('Tempo limite excedido ao esperar pelo botão "Próxima". Finalizando.', log_filename)
                break

    except Exception as e:
        log(f'Erro crítico ocorrido durante a raspagem: {e}', log_filename)
    finally:
        input("Pressione Enter para fechar o navegador...")
        driver.quit()
        log('Navegador fechado. Processo de raspagem concluído.', log_filename)

if __name__ == "__main__":
    raspar_comprasnet()